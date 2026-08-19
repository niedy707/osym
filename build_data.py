# -*- coding: utf-8 -*-
import json, os, re
from rank_model import rank, NTOT

ILLER = """ADANA ADIYAMAN AFYONKARAHİSAR AĞRI AMASYA ANKARA ANTALYA ARTVİN AYDIN BALIKESİR
BİLECİK BİNGÖL BİTLİS BOLU BURDUR BURSA ÇANAKKALE ÇANKIRI ÇORUM DENİZLİ DİYARBAKIR EDİRNE
ELAZIĞ ERZİNCAN ERZURUM ESKİŞEHİR GAZİANTEP GİRESUN GÜMÜŞHANE HAKKARİ HATAY ISPARTA MERSİN
İSTANBUL İZMİR KARS KASTAMONU KAYSERİ KIRKLARELİ KIRŞEHİR KOCAELİ KONYA KÜTAHYA MALATYA
MANİSA KAHRAMANMARAŞ MARDİN MUĞLA MUŞ NEVŞEHİR NİĞDE ORDU RİZE SAKARYA SAMSUN SİİRT SİNOP
SİVAS TEKİRDAĞ TOKAT TRABZON TUNCELİ ŞANLIURFA UŞAK VAN YOZGAT ZONGULDAK AKSARAY BAYBURT
KARAMAN KIRIKKALE BATMAN ŞIRNAK BARTIN ARDAHAN IĞDIR YALOVA KARABÜK KİLİS OSMANİYE DÜZCE
LEFKOŞA GAZİMAĞUSA GİRNE GÜZELYURT KKTC""".split()

MANUEL = {
    'GEBZE TEKNİK ÜNİVERSİTESİ': 'KOCAELİ',
    'TÜRK-JAPON BİLİM VE TEKNOLOJİ ÜNİVERSİTESİ': 'İSTANBUL',
}

def sehir(uni):
    if uni in MANUEL: return MANUEL[uni]
    if re.search(r'\((?:[^)]*-)?(AZERBAYCAN|KAZAKİSTAN|KIRGIZİSTAN|ARNAVUTLUK|MAKEDONYA|BOSNA[^)]*|SARAYBOSNA[^)]*)\)', uni):
        return 'YURT DIŞI'
    m = re.findall(r'\(([^)]*)\)', uni)
    for g in m:
        g = g.strip()
        if g in ILLER: return g
        for il in ILLER:
            if il in g: return il
    for il in sorted(ILLER, key=len, reverse=True):
        if uni.startswith(il) or f' {il}' in uni: return il
    return ''

def parse(pa):
    """Program adindan temel ad + nitelikleri cikar."""
    parens = re.findall(r'\(([^)]*)\)', pa)
    base = re.sub(r'\s*\([^)]*\)', '', pa).strip()
    low = pa.lower()
    if '%50' in pa: burs = '%50 İndirimli'
    elif '%25' in pa: burs = '%25 İndirimli'
    elif '%75' in pa: burs = '%75 İndirimli'
    elif 'burslu' in low: burs = 'Burslu'
    elif 'ücretli' in low: burs = 'Ücretli'
    else: burs = '—'
    if 'ingilizce' in low: dil = 'İngilizce'
    elif 'almanca' in low: dil = 'Almanca'
    elif 'fransızca' in low: dil = 'Fransızca'
    elif 'arapça' in low: dil = 'Arapça'
    elif 'rusça' in low: dil = 'Rusça'
    else: dil = 'Türkçe'
    return base, burs, dil, parens

def num(s):
    s = (s or '').strip()
    if s in ('', '--'): return None
    try: return float(s)
    except: return None

def i(s):
    v = num(s)
    return int(v) if v is not None else 0

QUOTAS = [('genel','Genel'), ('ob','Okul Birincisi'), ('y34','34 Yaş Üstü Kadın'), ('sehit','Şehit/Gazi Yakını')]

COLS = ["program_kodu","universite_turu","universite_adi","fakulte_adi","program_adi","puan_turu",
        "genel_kontenjan","genel_yerlesen","genel_min_puan","genel_max_puan",
        "ob_kontenjan","ob_yerlesen","ob_min_puan","ob_max_puan",
        "y34_kontenjan","y34_yerlesen","y34_min_puan","y34_max_puan",
        "sehit_kontenjan","sehit_yerlesen","sehit_min_puan","sehit_max_puan"]

def oku(path):
    """OSYM TABLO-3/TABLO-4 xlsx dosyasini sozluk listesine cevirir (ilk 3 satir baslik)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for r in ws.iter_rows(min_row=4, values_only=True):
        if r[0] is None or str(r[0]).strip() == '':
            continue
        yield {k: ('' if c is None else str(c).strip()) for k, c in zip(COLS, r[:22])}

def build(path, duzey):
    out = []
    for x in oku(path):
        uni = x['universite_adi']
        base, burs, dil, parens = parse(x['program_adi'])
        low = x['program_adi'].lower()
        mn, mx = num(x['genel_min_puan']), num(x['genel_max_puan'])
        pt = x['puan_turu']
        r_mn, c_mn = rank(mn, pt)
        r_mx, c_mx = rank(mx, pt)
        rec = {
            'kod': x['program_kodu'], 'uni': uni, 'unituru': x['universite_turu'],
            'fak': x['fakulte_adi'], 'prog': x['program_adi'], 'base': base,
            'pt': pt, 'duzey': duzey, 'sehir': sehir(uni), 'burs': burs, 'dil': dil,
            'kont': i(x['genel_kontenjan']), 'yer': i(x['genel_yerlesen']),
            'min': mn, 'max': mx, 'smin': r_mn, 'smax': r_mx,
            # yuzde: adayin KENDI puan turu icindeki yuzdelik dilimi.
            # Farkli puan turleri ancak bu deger uzerinden karsilastirilabilir
            # (SAY'da 1.135.718 aday varken DIL'de 132.826 aday var).
            'yuzde': (round(100 * r_mn / NTOT[pt], 4)
                      if (r_mn is not None and pt in NTOT) else None),
            'guven': c_mn or c_mx or '',
            'kktc': 'kktc uyruklu' in low, 'uolp': 'uolp' in low,
            'uzaktan': 'uzaktan' in low, 'io': 'ikinci öğretim' in low,
            'aof': 'açıköğretim' in low,
        }
        rec['bos'] = rec['kont'] - rec['yer']
        rec['kktc_uni'] = rec['unituru'] in ('KKTC', 'YURTDISI VAKIF')
        for k, lbl in QUOTAS[1:]:
            rec[k + '_k'] = i(x[k + '_kontenjan']); rec[k + '_y'] = i(x[k + '_yerlesen'])
            rec[k + '_min'] = num(x[k + '_min_puan']); rec[k + '_max'] = num(x[k + '_max_puan'])
        rec['tk'] = rec['kont'] + rec['ob_k'] + rec['y34_k'] + rec['sehit_k']
        rec['ty'] = rec['yer'] + rec['ob_y'] + rec['y34_y'] + rec['sehit_y']
        # acik = kontenjan turu bazinda POZITIF bosluklarin toplami
        # fazla = esit puan nedeniyle kontenjan ustu yerlestirme
        acik = fazla = 0
        for kk, yy in ((rec['kont'], rec['yer']), (rec['ob_k'], rec['ob_y']),
                       (rec['y34_k'], rec['y34_y']), (rec['sehit_k'], rec['sehit_y'])):
            d = kk - yy
            if d > 0: acik += d
            elif d < 0: fazla += -d
        rec['acik'] = acik
        rec['fazla'] = fazla
        out.append(rec)
    return out

def kaynak_dogrula():
    """kaynak/SHA256SUMS ile ÖSYM dosyalarinin degismedigini teyit eder.

    Projenin tum guvenilirlik iddiasi "bu sayilar ÖSYM'nin resmi dosyasindan
    birebir geldi" cumlesine dayaniyor; bozulmus ya da yanlislikla degistirilmis
    bir kaynak dosyasiyla veri uretilmesini engeller.
    """
    import hashlib
    yol = 'kaynak/SHA256SUMS'
    if not os.path.exists(yol):
        print('UYARI: kaynak/SHA256SUMS yok, butunluk dogrulamasi atlandi')
        return
    hata = []
    for satir in open(yol, encoding='utf-8'):
        satir = satir.strip()
        if not satir:
            continue
        beklenen, ad = satir.split()[0], satir.split()[-1]
        p = os.path.join('kaynak', ad)
        if not os.path.exists(p):
            hata.append(f'{ad}: dosya yok')
            continue
        h = hashlib.sha256(open(p, 'rb').read()).hexdigest()
        if h != beklenen:
            hata.append(f'{ad}: ozet uyusmuyor\n    beklenen {beklenen}\n    bulunan  {h}')
    if hata:
        raise SystemExit('KAYNAK DOGRULAMA BASARISIZ:\n  ' + '\n  '.join(hata))
    print('kaynak dosyalari dogrulandi (SHA-256)')


kaynak_dogrula()
rows = build('kaynak/tablo4.xlsx', 'Lisans') + build('kaynak/tablo3.xlsx', 'Ön Lisans')
print('Toplam program:', len(rows))
os.makedirs('data', exist_ok=True)
json.dump(rows, open('data/programlar.json', 'w', encoding='utf-8'), ensure_ascii=False, separators=(',', ':'))
print('JSON boyutu: %.1f MB' % (os.path.getsize('data/programlar.json')/1e6))

# ozet
tip = [r for r in rows if r['base'] == 'Tıp' and r['duzey'] == 'Lisans']
print('Tıp programı:', len(tip), '| toplam kontenjan:', sum(r['tk'] for r in tip),
      '| yerleşen:', sum(r['ty'] for r in tip))
print('Şehri boş kalan üni sayısı:', len({r['uni'] for r in rows if not r['sehir']}))
